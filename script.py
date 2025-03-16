import os
import requests
import openpyxl
from openpyxl.drawing.image import Image
from tkinter import filedialog, Tk, simpledialog

def download_image(url, save_path):
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, 'wb') as f:
            f.write(response.content)
        return save_path
    return None

def process_excel():
    root = Tk()
    root.withdraw()  # Hide the root window

    # Select Excel file
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        print("No file selected.")
        return

    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Display column names with numbers
    column_names = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    column_options = "\n".join([f"{i + 1}: {col_name}" for i, col_name in enumerate(column_names)])

    # Ask user to select column by number
    col_number = simpledialog.askinteger("Select Column", f"Select the column number for image URLs:\n\n{column_options}\n\nEnter a number:")
    if not col_number or col_number < 1 or col_number > ws.max_column:
        print("Invalid column number selected.")
        return

    # Find the last column and insert a new column for images
    last_col = ws.max_column + 1
    new_col_letter = openpyxl.utils.get_column_letter(last_col)
    ws.cell(row=1, column=last_col, value="Downloaded Image")  # Add header

    # Create directory for downloaded images
    image_dir = os.path.join(os.path.dirname(file_path), "downloaded_images")
    os.makedirs(image_dir, exist_ok=True)

    # Get all URLs from the selected column
    urls = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_number, max_col=col_number):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                urls.append(cell.value)

    print(f"Total images to download: {len(urls)}")

    # Download and insert images into the new column
    for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        url = row[col_number - 1].value
        if url and isinstance(url, str) and url.startswith("http"):
            image_filename = f"image_{index}.jpg"
            image_path = os.path.join(image_dir, image_filename)
            download_image(url, image_path)

            print(f"Downloaded {index - 1}/{len(urls)}")

            # Insert image into the new column
            img = Image(image_path)
            ws.add_image(img, f"{new_col_letter}{index}")

    # Save the updated file
    output_path = file_path.replace(".xlsx", "_with_images.xlsx")
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")

if __name__ == "__main__":
    process_excel()
